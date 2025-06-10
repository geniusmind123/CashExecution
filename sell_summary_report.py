import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import PatternFill

def parse_order_time(ts_str):
    """Parse order timestamp using multiple potential formats."""
    for fmt in (
        '%Y-%m-%d %H:%M:%S.%f',  # With microseconds
        '%Y-%m-%d %H:%M:%S',     # Without microseconds
        '%d-%m-%Y %H:%M:%S',     # Day-month-year format
        '%m/%d/%Y %H:%M:%S',     # Month/day/year
        '%d-%b-%Y %H:%M:%S',     # With abbreviated month (e.g., 13-Feb-2024)
    ):
        try:
            return datetime.strptime(ts_str, fmt)
        except ValueError:
            continue
    print(f"Warning: Failed to parse timestamp '{ts_str}'")
    return None

def generate_summary_data_sell(client_details, order_start_time, stock_details, client_apis):
    """
    Generate summary data for each client's sell operations using live positions and order book details.
    
    Args:
        client_details (dict): Contains each client's sell configurations including target quantities.
        order_start_time (datetime): The time after which orders are considered.
        stock_details (dict): Global dictionary containing stock instrument details, keyed by instrument id.
        client_apis (dict): Dictionary mapping client IDs to their API instances.
    
    Returns:
        dict: Summary data with structure {client_id: [{"Instrument": str, "Target": int, ...}, ...]}
    """
    summary_data = {}
    
    for client_id, params in client_details.items():
        client_summary = []
        api = client_apis.get(client_id)
        if not api:
            print(f"API not found for client {client_id}")
            continue
        
        # Fetch latest data
        try:
            order_book_response = api.get_order_book(clientID=client_id)
            pos_response = api.get_position_daywise(clientID=client_id)
            holding_response = api.get_holding(clientID=client_id)
        except Exception as e:
            print(f"Error fetching data for {client_id}: {str(e)}")
            continue
        
        # Process live positions
        live_positions = {}
        if pos_response.get('type') == 'success':
            for pos in pos_response['result'].get('positionList', []):
                inst_id = (pos.get('ExchangeInstrumentID') or 
                          pos.get('ExchangeNSEInstrumentID') or 
                          pos.get('ExchangeNSEInstrumentId'))
                if inst_id:
                    try:
                        inst_id = int(inst_id)
                        live_positions[inst_id] = {
                            'trading_symbol': pos.get('TradingSymbol', 'N/A'),
                            'quantity': pos.get('Quantity', 0)
                        }
                    except Exception as e:
                        print(f"Error processing position for {client_id}: {e}")
        
        # Process holdings
        holdings = {}
        if holding_response.get('type') == 'success':
            for isin, holding in holding_response['result']['RMSHoldings']['Holdings'].items():
                inst_id = (holding.get('ExchangeNSEInstrumentID') or 
                          holding.get('ExchangeNSEInstrumentId'))
                if inst_id:
                    try:
                        inst_id = int(inst_id)
                        holdings[inst_id] = {
                            'quantity': holding.get('HoldingQuantity', 0),
                            'buy_avg_price': holding.get('BuyAvgPrice', 0)
                        }
                    except Exception as e:
                        print(f"Error processing holding for {client_id}: {e}")
        
        # Process orders
        instrument_orders = defaultdict(lambda: {'executed': 0, 'open': 0, 'rejected': 0})
        orders = order_book_response.get('result', []) if order_book_response.get('type') == 'success' else []
        
        for order in orders:
            try:
                order_time = parse_order_time(order.get('OrderDateTime', ''))
                if not order_time or order_time < order_start_time:
                    continue
                
                inst_id = int(order.get('ExchangeInstrumentID', 0))
                if not inst_id:
                    continue
                
                status = order.get('OrderStatus', '').upper()
                quantity = int(order.get('OrderQuantity', 0))
                
                if status == 'FILLED':
                    instrument_orders[inst_id]['executed'] += quantity
                elif status == 'OPEN':
                    instrument_orders[inst_id]['open'] += quantity
                elif status == 'REJECTED':
                    instrument_orders[inst_id]['rejected'] += quantity
            except Exception as e:
                print(f"Error processing order for {client_id}: {e}")
        
        # Build summary for each instrument
        for ex_id, target_qty in params.get('total_quantity_to_sell', {}).items():
            try:
                ex_id_int = int(ex_id)
                stock_name = stock_details.get(ex_id_int, {}).get('name', f"ID:{ex_id}")
                orders_data = instrument_orders.get(ex_id_int, {'executed': 0, 'open': 0, 'rejected': 0})
                live_qty = live_positions.get(ex_id_int, {}).get('quantity', 0)
                holding_qty = holdings.get(ex_id_int, {}).get('quantity', 0)
                
                client_summary.append({
                    "Instrument": stock_name,
                    "Target": target_qty,
                    "Executed": orders_data['executed'],
                    "Open": orders_data['open'],
                    "Rejected": orders_data['rejected'],
                    "Live Pos": live_qty,
                    "Holding": holding_qty,
                    "Remaining": holding_qty - orders_data['executed']
                })
            except Exception as e:
                print(f"Error processing instrument {ex_id} for {client_id}: {e}")
        
        summary_data[client_id] = client_summary
    
    return summary_data

def export_summary_to_excel_sell(summary_data, filename="trading_summary_sell.xlsx"):
    """
    Export sell summary data to an Excel file with one sheet per client.
    
    Args:
        summary_data (dict): Dictionary with client IDs as keys and lists of instrument details as values.
        filename (str): Name of the Excel file to create.
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for client, instruments in summary_data.items():
            df = pd.DataFrame(instruments)
            desired_columns = ['Instrument', 'Target', 'Executed', 'Open', 'Rejected', 'Live Pos', 'Holding', 'Remaining']
            df = df[[col for col in desired_columns if col in df.columns]]
            
            # Add summary statistics
            total_row = pd.DataFrame([{
                'Instrument': 'TOTAL',
                'Target': df['Target'].sum(),
                'Executed': df['Executed'].sum(),
                'Open': df['Open'].sum(),
                'Rejected': df['Rejected'].sum(),
                'Live Pos': df['Live Pos'].sum(),
                'Holding': df['Holding'].sum(),
                'Remaining': df['Remaining'].sum()
            }])
            df = pd.concat([df, total_row], ignore_index=True)
            
            # Write to Excel with formatting
            df.to_excel(writer, sheet_name=client[:31], index=False)
            
            # Get the worksheet for formatting
            worksheet = writer.sheets[client[:31]]
            
            # Format the header row
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Format the total row
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=len(df), column=col_num)
                cell.font = cell.font.copy(bold=True)
                cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    print(f"Excel report generated: {filename}")

def generate_excel_summary_sell(client_details, order_start_time, stock_details, client_apis, filename="trading_summary_sell.xlsx"):
    """Generate and export the Excel summary for sell operations."""
    summary_data = generate_summary_data_sell(client_details, order_start_time, stock_details, client_apis)
    export_summary_to_excel_sell(summary_data, filename) 